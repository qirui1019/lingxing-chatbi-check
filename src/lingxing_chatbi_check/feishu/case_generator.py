from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re
from typing import Any

import yaml
from openpyxl import load_workbook


HELPER_TOOLS = {
    "ad_auth_shops",
    "get_my_sids",
    "erp_listing",
    "get_multi_platform_shop_list",
}

HEADER_FIELDS = [
    "category",
    "tool",
    "description",
    "required_arguments",
    "optional_arguments",
    "test_arguments",
    "output_fields",
    "dimensions",
    "database_table",
    "database_fields",
]

DATE_DIMENSIONS = {"report_date"}
FILTER_ONLY_DIMENSIONS = {"sponsored_type"}
NON_NUMERIC_FIELDS = {
    "rank_category",
    "small_cate_rank",
    "small_category_rank_json",
}
LATEST_DAY_DATABASE_FIELDS = {
    "chatbi.sale_report_asin": {
        "cate_rank",
        "avg_star",
        "reviews_count",
    },
}
TABLE_METRIC_MAPPINGS = {
    "chatbi.sale_report_msku_order": {
        "volume": "order_units",
        "amount": "order_sales_amount",
        "promotion_discount": "order_promotion_discount",
        "net_amount": "order_net_sales_amount",
        "gross_profit": "order_gross_profit",
        "impressions": "ad_impressions",
        "clicks": "ad_clicks",
        "spend": "ad_cost",
        "ad_order_quantity": "ad_order_count",
        "ads_sp_cost": "sp_ad_cost",
        "shared_ads_sb_cost": "sb_ad_cost",
        "shared_ads_sbv_cost": "sbv_ad_cost",
        "ads_sd_cost": "sd_ad_cost",
        "ads_sp_sales": "sp_ad_sales_amount",
        "shared_ads_sb_sales": "sb_ad_sales_amount",
        "shared_ads_sbv_sales": "sbv_ad_sales_amount",
        "ads_sd_sales": "sd_ad_sales_amount",
    },
    "chatbi.sale_report_asin": {
        "sessions": "sessions_browser",
        "page_views": "page_views_browser",
    },
    "chatbi.fba_list": {
        "real_transit_quantity": "afn_erp_real_shipped_quantity",
    },
    "chatbi.sp_keyword_report": {
        "spends": "cost",
        "direct_orders": "same_orders",
        "direct_sales": "same_sales",
        "ad_units": "units",
        "direct_units": "same_units",
    },
    "chatbi.sp_target_report": {
        "spends": "cost",
        "direct_orders": "same_orders",
        "direct_sales": "same_sales",
        "ad_units": "units",
        "direct_units": "same_units",
    },
    "chatbi.sp_search_term_report": {
        "spends": "cost",
        "direct_orders": "same_orders",
        "direct_sales": "same_sales",
        "ad_units": "units",
        "direct_units": "same_units",
    },
}


@dataclass(frozen=True)
class AvailableToolRow:
    category: str
    tool: str
    required_arguments: str
    optional_arguments: str
    test_arguments: str
    output_fields: str
    dimensions: str
    database_table: str
    database_fields: str


def load_available_tool_rows(path: Path) -> list[AvailableToolRow]:
    workbook = load_workbook(path, data_only=True)
    sheet_name = "可用tool" if "可用tool" in workbook.sheetnames else workbook.sheetnames[0]
    sheet = workbook[sheet_name]

    rows: list[AvailableToolRow] = []
    last: dict[str, Any] = {}
    for row_index in range(2, sheet.max_row + 1):
        row: dict[str, Any] = {
            field: sheet.cell(row_index, column).value
            for column, field in enumerate(HEADER_FIELDS, start=1)
        }
        if not any(value is not None for value in row.values()):
            continue

        for key in ("category", "tool", "required_arguments", "optional_arguments"):
            if row.get(key) in (None, "") and last.get(key) not in (None, ""):
                row[key] = last[key]
        last.update({key: value for key, value in row.items() if value not in (None, "")})

        tool = str(row.get("tool") or "").strip()
        database_table = _normalize_database_table(str(row.get("database_table") or ""))
        if not tool or not database_table or database_table.startswith("tool返回"):
            continue

        rows.append(
            AvailableToolRow(
                category=str(row.get("category") or ""),
                tool=tool,
                required_arguments=str(row.get("required_arguments") or ""),
                optional_arguments=str(row.get("optional_arguments") or ""),
                test_arguments=str(row.get("test_arguments") or ""),
                output_fields=str(row.get("output_fields") or ""),
                dimensions=str(row.get("dimensions") or ""),
                database_table=database_table,
                database_fields=str(row.get("database_fields") or ""),
            )
        )
    return rows


def write_case_templates(path: Path, output_dir: Path) -> list[Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    written: list[Path] = []
    for row in load_available_tool_rows(path):
        if row.tool in HELPER_TOOLS:
            continue
        case_data = build_case_template(row)
        filename = f"{row.tool}__{row.database_table}.yml".replace("/", "_")
        target = output_dir / filename
        target.write_text(
            yaml.safe_dump(case_data, allow_unicode=True, sort_keys=False),
            encoding="utf-8",
        )
        written.append(target)
    return written


def build_case_template(row: AvailableToolRow) -> dict[str, Any]:
    arguments = _parse_arguments(row.test_arguments)
    if "sponsored_type=sp" in row.dimensions:
        arguments.setdefault("sponsored_type", "sp")

    raw_dimensions = _split_dimensions(row.dimensions)
    dimensions = _compare_dimensions_for(row)
    database_fields = _numeric_database_fields(row)
    tool_fields = _numeric_tool_fields(row)

    return {
        "enabled": False,
        "name": f"{row.tool} vs {row.database_table}",
        "auth": {"mode": "all_users"},
        "scope": {
            "shop_discovery": _shop_discovery_for(row),
            "listing_mapping": "erp_listing"
            if any(key in row.dimensions.lower() for key in ("msku", "asin", "fnsku"))
            else None,
        },
        "tool": {
            "name": row.tool,
            "arguments": arguments,
            "dynamic_arguments": _dynamic_arguments_for(row),
            "pagination": _pagination_for(row),
        },
        "database": {
            "table": row.database_table,
            "sql": _sql_template_for(row),
            "params": _database_params_for(arguments),
        },
        "compare": {
            "dimensions": dimensions,
            "metrics": database_fields,
            "dimension_mappings": _dimension_mappings_for(row),
            "metric_mappings": _metric_mappings(tool_fields, database_fields, row.database_table),
            "tolerance": 0.01,
        },
        "notes": {
            "raw_dimensions": _split_lines(row.dimensions, split_plus=False),
            "inactive_dimensions": _inactive_dimensions_for(raw_dimensions, dimensions),
            "tool_output_fields": _split_fields(row.output_fields),
            "non_numeric_fields": _non_numeric_fields(row),
            "source": "Generated from data/feishu/lingxing_mcp_tools.xlsx; review SQL and field paths before enabling.",
        },
    }


def _dynamic_arguments_for(row: AvailableToolRow) -> dict[str, Any]:
    if _is_ad_category(row.category):
        return {
            "shop_argument": "profile_ids",
            "shop_batch_mode": "list",
            "source_field": "profile_id",
            "database_source_field": "sid",
            "batch_size": 3,
            "database_param": "sid_values",
        }
    if row.tool == "query_product_performance_asin_lists":
        return {
            "shop_argument": "sids",
            "shop_batch_mode": "list",
            "source_field": "sid",
            "batch_size": 2,
            "database_param": "sid_values",
        }
    if "sid=" in row.test_arguments:
        return {
            "shop_argument": "sid",
            "shop_batch_mode": "single",
            "source_field": "sid",
            "batch_size": 1,
            "database_param": "sid_values",
        }
    if "sids" in row.optional_arguments:
        return {
            "shop_argument": "sids",
            "shop_batch_mode": "list",
            "source_field": "sid",
            "batch_size": 50,
            "database_param": "sid_values",
        }
    return {
        "shop_argument": "sid",
        "shop_batch_mode": "single",
        "source_field": "sid",
        "batch_size": 1,
        "database_param": "sid_values",
    }


def _pagination_for(row: AvailableToolRow) -> dict[str, Any] | None:
    if row.tool == "query_product_performance_asin_lists":
        return {
            "enabled": True,
            "page_argument": "offset",
            "page_start": 0,
            "page_size_argument": "length",
            "page_size": 1000,
            "max_pages": 1000,
            "page_value_mode": "offset",
        }

    optional_arguments = set(_split_fields(row.optional_arguments))
    if {"page", "length"}.issubset(optional_arguments):
        return {
            "enabled": True,
            "page_argument": "page",
            "page_start": 1,
            "page_size_argument": "length",
            "page_size": 500 if _is_ad_category(row.category) else 1000,
            "max_pages": 1000,
        }
    return None


def _shop_discovery_for(row: AvailableToolRow) -> str:
    if _is_ad_category(row.category):
        return "ad_auth_shops"
    return "get_my_sids"


def _parse_arguments(text: str) -> dict[str, Any]:
    arguments: dict[str, Any] = {}
    for line in _split_lines(text):
        if "=" not in line:
            continue
        key, value = line.split("=", 1)
        clean_key = key.strip()
        if clean_key in {"sid", "sids", "profile_id", "profile_ids"}:
            continue
        clean_value = re.split(r"[（(]", value, maxsplit=1)[0].strip()
        arguments[clean_key] = _parse_value(clean_value)
    return arguments


def _parse_value(value: str) -> Any:
    if value in ("-", ""):
        return ""
    if value.startswith("[") and value.endswith("]"):
        try:
            return yaml.safe_load(value)
        except yaml.YAMLError:
            return value
    if value.isdigit():
        return int(value)
    return value


def _database_params_for(arguments: dict[str, Any]) -> dict[str, Any]:
    params = {
        key: value
        for key, value in arguments.items()
        if key in {"start_date", "end_date", "report_date"}
    }
    report_date = params.get("report_date")
    if "report_date" not in params and "start_date" in params and "end_date" in params:
        params["report_date"] = f"{params['start_date']} - {params['end_date']}"
        report_date = params["report_date"]
    if isinstance(report_date, str):
        match = re.match(
            r"\s*(\d{4}-\d{2}-\d{2})\s*-\s*(\d{4}-\d{2}-\d{2})\s*$",
            report_date,
        )
        if match:
            params["report_start_date"] = match.group(1)
            params["report_end_date"] = match.group(2)
    return params


def _sql_template_for(row: AvailableToolRow) -> str:
    dimensions = _compare_dimensions_for(row)
    group_dimensions = _sql_group_dimensions_for(row, dimensions)
    fields = _numeric_database_fields(row)
    if row.database_table in LATEST_DAY_DATABASE_FIELDS:
        return _latest_day_sql_template_for(
            row,
            dimensions=dimensions,
            group_dimensions=group_dimensions,
            fields=fields,
        )

    selected = _sql_dimension_selects_for(row, dimensions) + [
        f"sum({field}) as {field}" for field in fields
    ]
    if not selected:
        selected = ["*"]
    where_clauses = ["sid in :sid_values"]
    arguments = _parse_arguments(row.test_arguments)
    if "start_date" in arguments and "end_date" in arguments:
        where_clauses.append("report_date between :start_date and :end_date")
    elif "report_date" in arguments:
        where_clauses.append("report_date between :report_start_date and :report_end_date")
    return (
        "select\n  "
        + ",\n  ".join(selected)
        + f"\nfrom {row.database_table}\nwhere "
        + "\n  and ".join(where_clauses)
        + _group_by_clause(group_dimensions)
    )


def _latest_day_sql_template_for(
    row: AvailableToolRow,
    *,
    dimensions: list[str],
    group_dimensions: list[str],
    fields: list[str],
) -> str:
    latest_fields = [
        field
        for field in fields
        if field in LATEST_DAY_DATABASE_FIELDS[row.database_table]
    ]
    sum_fields = [field for field in fields if field not in set(latest_fields)]
    where_clauses = _where_clauses_for(row)
    ctes = [
        (
            "base as (\n"
            f"  select *\n  from {row.database_table}\n"
            "  where "
            + "\n    and ".join(where_clauses)
            + "\n)"
        )
    ]
    if latest_fields:
        ctes.append(
            "latest as (\n"
            "  select\n"
            "    sid,\n"
            "    asin,\n"
            + ",\n".join(f"    {field}" for field in latest_fields)
            + "\n"
            "  from (\n"
            "    select\n"
            "      sid,\n"
            "      asin,\n"
            + ",\n".join(f"      {field}" for field in latest_fields)
            + ",\n"
            "      row_number() over (partition by sid, asin order by report_date desc) as rn\n"
            "    from base\n"
            "  ) ranked\n"
            "  where rn = 1\n"
            ")"
        )
    if sum_fields:
        ctes.append(
            "agg as (\n"
            "  select\n"
            "    "
            + ",\n    ".join(
                group_dimensions + [f"sum({field}) as {field}" for field in sum_fields]
            )
            + "\n"
            "  from base"
            + _group_by_clause(group_dimensions)
            + ")"
        )

    source = "latest" if latest_fields else "agg"
    select_columns = _latest_day_select_columns(
        dimensions=dimensions,
        latest_fields=latest_fields,
        sum_fields=sum_fields,
        source=source,
    )
    sql = "with\n" + ",\n".join(ctes) + "\nselect\n  " + ",\n  ".join(select_columns)
    sql += f"\nfrom {source}"
    if latest_fields and sum_fields:
        sql += "\nleft join agg on latest.sid = agg.sid and latest.asin = agg.asin"
    return sql + "\n"


def _latest_day_select_columns(
    *,
    dimensions: list[str],
    latest_fields: list[str],
    sum_fields: list[str],
    source: str,
) -> list[str]:
    selected: list[str] = []
    for dimension in dimensions:
        if dimension == "report_date":
            selected.append(":report_date as report_date")
        else:
            selected.append(f"{source}.{dimension} as {dimension}")
    selected.extend(f"latest.{field} as {field}" for field in latest_fields)
    selected.extend(f"agg.{field} as {field}" for field in sum_fields)
    return selected or ["*"]


def _where_clauses_for(row: AvailableToolRow) -> list[str]:
    where_clauses = ["sid in :sid_values"]
    arguments = _parse_arguments(row.test_arguments)
    if "start_date" in arguments and "end_date" in arguments:
        where_clauses.append("report_date between :start_date and :end_date")
    elif "report_date" in arguments:
        where_clauses.append("report_date between :report_start_date and :report_end_date")
    return where_clauses


def _group_by_clause(dimensions: list[str]) -> str:
    if not dimensions:
        return "\n"
    return "\ngroup by\n  " + ",\n  ".join(dimensions) + "\n"


def _split_dimensions(text: str) -> list[str]:
    fields: list[str] = []
    for line in _split_lines(text, split_plus=True):
        clean_line = _strip_parenthetical_notes(line)
        field = _first_identifier(clean_line)
        mapped = _extract_lingxing_field(line)
        if field == "keyword_id" and mapped == "target_id":
            field = "target_id"
        if field and field not in DATE_DIMENSIONS | FILTER_ONLY_DIMENSIONS:
            fields.append(field)
    return fields


def _compare_dimensions_for(row: AvailableToolRow) -> list[str]:
    if _is_ad_category(row.category):
        return ["sid", "report_date"]
    dimensions = _split_dimensions(row.dimensions)
    if (
        row.tool == "query_product_performance_asin_lists"
        and "report_date" in str(row.dimensions)
        and "report_date" not in dimensions
    ):
        dimensions.append("report_date")
    return dimensions


def _sql_dimension_selects_for(
    row: AvailableToolRow,
    dimensions: list[str],
) -> list[str]:
    if _is_ad_category(row.category) or row.tool == "query_product_performance_asin_lists":
        return [
            ":report_date as report_date" if dimension == "report_date" else dimension
            for dimension in dimensions
        ]
    return dimensions


def _sql_group_dimensions_for(
    row: AvailableToolRow,
    dimensions: list[str],
) -> list[str]:
    if _is_ad_category(row.category) or row.tool == "query_product_performance_asin_lists":
        return [dimension for dimension in dimensions if dimension != "report_date"]
    return dimensions


def _dimension_mappings_for(row: AvailableToolRow) -> dict[str, str]:
    if _is_ad_category(row.category):
        return {"sid": "sid", "report_date": "report_date"}
    mappings = _dimension_mappings(row.dimensions, row.category)
    if (
        row.tool == "query_product_performance_asin_lists"
        and "report_date" in str(row.dimensions)
    ):
        mappings["report_date"] = "report_date"
    return mappings


def _inactive_dimensions_for(
    raw_dimensions: list[str],
    active_dimensions: list[str],
) -> list[str]:
    active = set(active_dimensions)
    return [dimension for dimension in raw_dimensions if dimension not in active]


def _split_fields(text: str) -> list[str]:
    fields: list[str] = []
    for line in _split_lines(text, split_plus=False):
        fields.extend(re.findall(r"[A-Za-z_][A-Za-z0-9_]*", _strip_parenthetical_notes(line)))
    return fields


def _split_lines(
    text: str,
    *,
    split_plus: bool = True,
    identifiers_only: bool = False,
) -> list[str]:
    values: list[str] = []
    normalized = str(text or "")
    if split_plus:
        normalized = _normalize_dimension_separators(normalized)
    for raw_line in normalized.splitlines():
        line = raw_line.strip()
        if not line or line in {"/", "-"}:
            continue
        if identifiers_only:
            identifier = _first_identifier(line)
            if identifier:
                values.append(identifier)
        else:
            values.append(line)
    return values


def _normalize_database_table(value: str) -> str:
    return value.strip().replace("cahtbi.", "chatbi.")


def _normalize_dimension_separators(value: str) -> str:
    return re.sub(r"(?m)^\s*\+", "", value)


def _ordered_mapping(tool_fields: list[str], database_fields: list[str]) -> dict[str, str]:
    return {
        tool_field: database_field
        for tool_field, database_field in zip(tool_fields, database_fields)
    }


def _metric_mappings(
    tool_fields: list[str],
    database_fields: list[str],
    database_table: str,
) -> dict[str, str]:
    overrides = TABLE_METRIC_MAPPINGS.get(database_table, {})
    mappings: dict[str, str] = {}
    used_tool_fields: set[str] = set()

    for database_field in database_fields:
        tool_field = _tool_field_for_database_field(
            database_field,
            tool_fields,
            overrides,
        )
        if tool_field is None:
            continue
        mappings[tool_field] = database_field
        used_tool_fields.add(tool_field)

    remaining_tool_fields = [
        field for field in tool_fields if field not in used_tool_fields
    ]
    remaining_database_fields = [
        field for field in database_fields if field not in set(mappings.values())
    ]
    mappings.update(_ordered_mapping(remaining_tool_fields, remaining_database_fields))
    return mappings


def _tool_field_for_database_field(
    database_field: str,
    tool_fields: list[str],
    overrides: dict[str, str],
) -> str | None:
    for tool_field, mapped_database_field in overrides.items():
        if mapped_database_field == database_field and tool_field in tool_fields:
            return tool_field
    if database_field in tool_fields:
        return database_field
    return None


def _dimension_mappings(text: str, category: str = "") -> dict[str, str]:
    mappings: dict[str, str] = {}
    normalized = _normalize_dimension_separators(str(text or ""))
    for raw_line in normalized.splitlines():
        line = raw_line.strip()
        if not line or line in {"/", "-"}:
            continue
        db_field = _first_identifier(_strip_parenthetical_notes(line))
        if not db_field or db_field in DATE_DIMENSIONS | FILTER_ONLY_DIMENSIONS:
            continue
        if _is_ad_category(category) and db_field == "sid":
            tool_field = "sid"
        else:
            tool_field = _extract_lingxing_field(line) or db_field
        if tool_field == "keyword_id" and db_field == "target_id":
            tool_field = "target_id"
        if db_field == "keyword_id" and tool_field == "target_id":
            db_field = "target_id"
        mappings[tool_field] = db_field
    return mappings


def _extract_lingxing_field(value: str) -> str | None:
    note_match = re.search(r"(?:领星中是|棰嗘槦涓槸)\s*([^）)]*)", value)
    if not note_match:
        return None
    note = note_match.group(1).strip()
    if "+" in note or "/" in note:
        return None
    match = re.search(r"[A-Za-z_][A-Za-z0-9_]*", note)
    if not match:
        return None
    return match.group(0)


def _strip_parenthetical_notes(value: str) -> str:
    value = re.sub(r"（[^）]*）", "", value)
    return re.sub(r"\([^)]*\)", "", value)


def _first_identifier(value: str) -> str | None:
    value = _strip_parenthetical_notes(value)
    match = re.search(r"[A-Za-z_][A-Za-z0-9_]*", value)
    if not match:
        return None
    return match.group(0)


def _is_ad_category(category: str) -> bool:
    return category in {"广告", "骞垮憡"}


def _numeric_database_fields(row: AvailableToolRow) -> list[str]:
    return [
        field
        for field in _split_fields(row.database_fields)
        if field not in NON_NUMERIC_FIELDS
    ]


def _numeric_tool_fields(row: AvailableToolRow) -> list[str]:
    return [
        field
        for field in _split_fields(row.output_fields)
        if field not in NON_NUMERIC_FIELDS
    ]


def _non_numeric_fields(row: AvailableToolRow) -> list[str]:
    fields = _split_fields(row.output_fields) + _split_fields(row.database_fields)
    return list(
        dict.fromkeys(field for field in fields if field in NON_NUMERIC_FIELDS)
    )
